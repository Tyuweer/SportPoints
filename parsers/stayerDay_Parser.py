import re
from openpyxl import load_workbook

class StayerDay_Parser:
    def __init__(self):
        self.NON_ATHLETE_KEYWORDS = [
            'протокол', 'технических', 'результатов', 'место', 'разряд',
            'фамилия', 'имя', 'год', 'рожд', 'команда', 'результат',
            'норматив', 'очки', 'предв', 'финал', 'главный', 'судья',
            'секретарь', 'соревнований', 'федерация', 'министерство',
            'первенство', 'чемпионат', 'соревнования', 'протокол',
            'февраля', 'марта', 'апреля', 'мая', 'июня', 'июля',
            'августа', 'сентября', 'октября', 'ноября', 'декабря',
            'января', 'дистанция', 'дисциплина', 'зачет', 'этап'
        ]
        self.EVENT_KEYWORDS = ['плавание', 'ныряние', 'подводное', 'классическ', 'ласт']
        self.AGE_GROUPS = ['юниоры', 'юниорки', 'юноши', 'девушки', 'мужчины', 'женщины', 'мальчики', 'девочки']

    def parse(self, excel_path, is_manual=True):
        wb = load_workbook(excel_path, data_only=True)
        events = []
        current_event = None
        seen_events = set()

        for sheet in wb.worksheets:
            rows = sheet.iter_rows(values_only=True)
            for row in rows:
                # Преобразуем ячейки в строки, заменяя None на пустую строку
                line_parts = [str(cell).strip() if cell is not None else '' for cell in row]
                # Убираем полностью пустые строки
                if not any(line_parts):
                    continue

                # Собираем "строку" как в PDF-парсере для совместимости логики
                fake_line = ' '.join(part for part in line_parts if part)

                # Пропускаем строки с "в/к"
                if any(kw in fake_line.lower() for kw in ['в/к', 'в.к.', 'вк']):
                    continue

                # Проверяем, является ли это заголовком дисциплины
                if self.is_event_header(fake_line):
                    #if line in seen_events:
                           # continue
                    if current_event:
                        events.append(current_event)
                    current_event = {
                        "event_name": fake_line,
                        "results": []
                    }
                  # seen_events.add(line)
                    continue

                # Пропускаем служебные строки
                if not self.is_athlete_row(line_parts):
                    continue

                # Парсим запись спортсмена из реальных ячеек (не из склеенной строки!)
                record = self.parse_result_row_excel(line_parts, is_manual=is_manual)
                if record and current_event:
                    current_event["results"].append(record)

            # Завершаем последнее событие после листа
            if current_event:
                events.append(current_event)
                current_event = None

        return events

    def is_event_header(self, line: str) -> bool:
        line_lower = line.lower()
        has_keyword = any(kw in line_lower for kw in self.EVENT_KEYWORDS)
        has_age = any(ag in line_lower for ag in self.AGE_GROUPS)
        return has_keyword and has_age

    def is_athlete_row(self, parts: list) -> bool:
        if not parts or all(p == '' for p in parts):
            return False

        text_check = ' '.join(p for p in parts[:5] if p).lower()

        # Исключаем по ключевым словам
        for kw in self.NON_ATHLETE_KEYWORDS:
            if kw in text_check:
                return False

        # Исключаем даты вроде "26 февраля"
        date_pattern = r'\d{1,2}\s*(января|февраля|марта|апреля|мая|июня|июля|августа|сентября|октября|ноября|декабря)'
        if re.search(date_pattern, text_check, re.IGNORECASE):
            return False

        return True

    def parse_result_row_excel(self, cells: list, is_manual=True):
        # Убираем пустые ячейки справа и слева
        trimmed = [c for c in cells if c != '']
        if not trimmed:
            return None

        try:
            idx = 0

            # Место
            place = None
            if trimmed[0].isdigit():
                place = trimmed[0]
                idx = 1

            # Разряд
            rank = None
            if idx < len(trimmed):
                candidate = trimmed[idx]
                if candidate in ['I', 'II', 'III', '1', '2', '3', 'МС', 'КМС', 'ЗМС', 'МСМК', 'б/р', 'б\\р']:
                    rank = candidate
                    idx += 1
                elif candidate in ['I юн', 'II юн', 'III юн']:
                    rank = candidate
                    idx += 1

            # ФИО — может занимать 1–3 ячейки
            name_parts = []
            while idx < len(trimmed):
                cell = trimmed[idx]
                # Если похоже на год рождения
                if re.fullmatch(r'\d{4}', cell):
                    break
                # Если похоже на результат
                if re.match(r'\d{1,2}[,.:]\d{2}([,.:]\d{2})?$', cell) or cell in ['DNS', 'DSQ', 'DNF']:
                    break
                # Если содержит точку и похоже на команду (например, "г. Ачинск")
                if '.' in cell and len(cell) > 3:
                    break
                name_parts.append(cell)
                idx += 1

            if not name_parts:
                return None
            full_name = ' '.join(name_parts)

            # Год рождения
            birth_year = None
            if idx < len(trimmed) and re.fullmatch(r'\d{4}', trimmed[idx]):
                birth_year = trimmed[idx]
                idx += 1

            # Команда — до результата
            team_parts = []
            while idx < len(trimmed):
                cell = trimmed[idx]
                if re.match(r'\d{1,2}[,.:]\d{2}([,.:]\d{2})?$', cell) or cell in ['DNS', 'DSQ', 'DNF']:
                    break
                team_parts.append(cell)
                idx += 1

            team = ' '.join(team_parts).strip()

            # Результат
            result = None
            if idx < len(trimmed):
                token = trimmed[idx]
                if re.match(r'\d{1,2}[,.:]\d{2}([,.:]\d{2})?$', token):
                    result = token.replace('.', ',')  # Excel может использовать точку
                    idx += 1

            # Остальное — норматив и очки (обычно не используется в Excel, но оставим для совместимости)
            normative = None
            points = None
            # Можно расширить при необходимости

            return {
                "place": place,
                "rank": rank,
                "full_name": full_name,
                "birth_year": birth_year,
                "team": team,
                "result": result,
                "normative": normative,
                "points": points,
                "is_manual_timing": is_manual
            }

        except Exception as e:
            print(f"Ошибка парсинга строки Excel: {cells} — {e}")
            return None